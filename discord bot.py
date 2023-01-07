import discord
import time
import math
from openpyxl import load_workbook
import os
from dotenv import load_dotenv

load_dotenv()  # take environment variables from .env.

TOKEN = os.getenv("TOKEN")
CREATOR = os.getenv("CREATOR")
EXCEL = os.getenv("EXCEL")
wb = load_workbook(EXCEL)
ws = wb.active

intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)


@client.event
async def on_ready():
    print('We have logged in as {0.user}'.format(client))


@client.event
async def on_message(message):
    def da(seconds=int(time.time()) - 104400):
        if (math.ceil(seconds / 604800)) % 3 == 2:
            x = ((math.ceil(seconds / 604800) * 604800) - seconds) + 604800
        else:
            x = ((math.ceil(seconds / 604800) * 604800) - seconds)
        return x

    def t5_da(name, t5_in_infirm, current_lunite, current_crafting_time=0):
        original_lunite_per_second_to_craft = 0.8333
        calc_lprod = ws[f'B{buscarv(ws, name)[1]}'].value
        calc_cspd = ws[f'C{buscarv(ws, name)[1]}'].value

        healing_capacity = ((int(da()) * int(calc_lprod) / 3600) + int(current_lunite)) / 40
        real_capacity = int(healing_capacity) - int(t5_in_infirm)

        lunite_per_hour_crafting = float(original_lunite_per_second_to_craft) * (float(calc_cspd) / 100 + 1) * 3600
        effectiveness = int(lunite_per_hour_crafting) / int(
            calc_lprod)  # 0.25 (25%) of your lunite production is used in crafting
        offset = int(current_crafting_time) * int(lunite_per_hour_crafting) / 40

        return f'WITHOUT CRAFTING:\n' \
               f'You could heal up to {math.ceil(healing_capacity + offset)} troops before dragon arena, which means you can send {math.ceil(real_capacity + offset)} t5 with buffers or {math.ceil((real_capacity + offset) / 6 * 10)} without buffers' \
               f'\n\n' \
               f'WHILE CRAFTING:\n' \
               f'You could heal up to {math.ceil(healing_capacity * (1 - effectiveness) + offset)} troops before dragon arena, which means you can send {math.ceil(real_capacity * (1 - effectiveness) + offset)} t5 with buffers or {math.ceil((real_capacity * (1 - effectiveness) + offset) / 6 * 10)} without buffers\n'

    def set_user(name, Lprod, Cspd):
        name_coords = f'A{buscarv(ws, str(message.author))[1]}'
        Lprod_coords = f'B{buscarv(ws, str(message.author))[1]}'
        Cspd_coords = f'C{buscarv(ws, str(message.author))[1]}'

        ws[name_coords] = name
        ws[Lprod_coords] = Lprod
        ws[Cspd_coords] = Cspd


    def create_user(name, Lprod, Cspd):
        name_coords = f'A{ws.max_row + 1}'
        Lprod_coords = f'B{ws.max_row + 1}'
        Cspd_coords = f'C{ws.max_row + 1}'

        ws[name_coords] = name
        ws[Lprod_coords] = Lprod
        ws[Cspd_coords] = Cspd

    def buscarv(ws, search_string, column="A"):
        for row in range(1, ws.max_row + 1):
            coordinate = "{}{}".format(column, row)
            if ws[coordinate].value == search_string:
                return column, row
        return column, None

    msg = message.content
    if message.author == client.user:
        return

    if msg.startswith('$new'):
        try:
            users = []
            for i in range(2, ws.max_row + 1):
                users.append(ws[f'A{i}'].value)
            if str(message.author) not in users:
                create_user(str(message.author), msg.split()[1], msg.split()[2])
                wb.save(EXCEL)
                await message.channel.send(
                    f'Okay {message.author}, the info that I have of you is:\n'
                    f'{msg.split()[1]} lunite production\n'
                    f'{msg.split()[2]} crafting speed\n')
            else:
                await message.channel.send(
                    f'Are you sure you didn\'t already create a user? I have this info of you:\n'
                    f'{ws[f"B{buscarv(ws, str(message.author))[1]}"].value} lunite production\n'
                    f'{ws[f"C{buscarv(ws, str(message.author))[1]}"].value} crafting speed\n'
                )
        except Exception as error:
            await message.channel.send(error)

    if msg.startswith('$new_guide'):
        await message.channel.send(
            'The format should be something along the lines of \"$new x y\", being x your (hourly) lunite production and y your crafting speed (with crafting gear)\n'
            'For example, if you make 12300 lunite per hour and your crafting speed is 154.89%, your command should be \"$new 12300 154.89\"'
        )

    if msg.startswith('$set'):
        try:
            users = []
            for i in range(2, ws.max_row + 1):
                users.append(ws[f'A{i}'].value)
            if str(message.author) in users:
                set_user(str(message.author), msg.split()[1], msg.split()[2])
                wb.save(EXCEL)
                await message.channel.send(
                    f'Okay {message.author}, the new info that I have of you is:\n'
                    f'{msg.split()[1]} lunite production\n'
                    f'{msg.split()[2]} crafting speed\n')
            else:
                await message.channel.send(f'I don\'t have any data about you, please use command \"$new\"')
        except Exception as error:
            await message.channel.send(error)

    if msg.startswith('$t5da'):
        try:
            users = []
            for i in range(2, ws.max_row + 1):
                users.append(ws[f'A{i}'].value)
            if str(message.author) in users:
                await message.channel.send(t5_da(str(message.author), msg.split()[1], msg.split()[2]))
            else:
                await message.channel.send(f'I don\'t have any data about you, please use command \"$new\"')
        except Exception as error:
            await message.channel.send(error)

    if msg.startswith('$t5da_guide'):
        await message.channel.send(
            'The format should be something along the lines of \"$t5da x y z\", being:\n\n'
            'x the t5 in your infirmaries (that\'s not being healed already)\n'
            'y the current lunite outside your bags (you could theoretically also sum some extra lunite if you\'re willing to spend from your bags)\n'
            'z the time left for the current crafting to finish, in hours (if it\'s 3h 30m, type 3.5).\n'
            'The default value for z is 0, you could just type x and y in the command and it would calculate it as if you had to start crafting right now\n\n'
            'For example, if you have 1329 t5 in infirmaries, of which 20 are healing right now, 327800 lunite open and your current crafting will end in 21 hours and 25 minutes, type \"$t5da 1309 327800 21.5\"\n'
            'Please keep in mind that the calculations don\'t care about speed ups and assume you will never reach maximum lunite capacity and that all the t5 that you send will die, you can estimate your number best according to each circumstance'
            'Also, important to note that I don\'t know which timeslot will you play DA in, and therefore my calculations are for the beginning of the DA\'s day (GMT -5 00:00)'
        )

    if msg.startswith('$da'):
        days = math.floor(da() / 86400)
        hours = math.floor((da() - (days * 86400)) / 3600)
        minutes = math.floor((da() - (days * 86400) - (hours * 3600)) / 60)
        await message.channel.send(
            f'There are {days} days, {hours} hours and {minutes} minutes left for next dragon arena.\n'
            f'Please note that I don\'t know at what time will your guild be playing dragon arena, I\'m calculating for GMT -5 at 00:00, so your DA will at least be one hour after those calculations)')

    if msg.startswith('$help'):
        await message.channel.send(
            'These are the available commands:\n\n'
            '$new - Make the bot store your information\n'
            '$new_guide - Shows a guide of how to use the command $New\n'
            '$set - Update the information of you stored by me\n'
            '$t5da - Calculates how many t5 you can send to a rally to be able to participate in the next DA without using bag lunite\n'
            '$t5da_guide - Shows a guide of how to use the command $t5da, plus some extra info\n'
            '$da - Shows how much time is left until next dragon arena\n'
            '$help - Shows this message\n\n'
            f'If you need further assistance, please PM my creator {CREATOR}')


client.run(TOKEN)
